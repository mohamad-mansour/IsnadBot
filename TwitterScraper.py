import logging
from logging.handlers import RotatingFileHandler
import random
import time
from datetime import datetime, timedelta, timezone
import pytz
from typing import List
from fastapi import BackgroundTasks, FastAPI, HTTPException, File, Depends, UploadFile, Header, Query, Path
from account import Account
from scraper import Scraper
from fastapi.responses import JSONResponse, PlainTextResponse
from fastapi.background import BackgroundTasks
from sqlalchemy import create_engine, Column, Integer, String, DateTime, inspect, Boolean, desc
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.sql.expression import false

# configure the log format
formatter  = logging.Formatter('%(asctime)s - %(message)s')


handler = RotatingFileHandler('app.log', maxBytes=1024*1024*10, backupCount=5)
handler.setFormatter(formatter)
# set the logging level to INFO
logging.basicConfig(format='%(asctime)s %(message)s', level=logging.INFO)
logging.getLogger('sqlalchemy').setLevel(logging.ERROR)

# add the handler to the logger
logger = logging.getLogger(__name__)
logger.addHandler(handler)

description = """

## Isnad Scrap - Util API <img src=\'https://flagcdn.com/24x18/ps.png\'> 🔻🔻🔻


Isnad scrap is a powerful tool designed to streamline interactions with the Twitter platform automatically.

The following APIs provide various services for managing user cookies, handling target user IDs, and reading file contents.

**Key Features:**

- **Upload Main User Cookies:**
    Upload the main Twitter user cookies responsible for actions like replying to or liking tweets. 
    The uploaded file should follow the format: `auth_token_value|ct0_value|username`, with each record on a new line.

- **Upload Scrap Cookies:**
    Update the list of scrap Twitter user cookies, specifically used for searching recent tweets. 
    The file format should match that of main user cookies.

- **Upload Target User IDs:**
    Maintain a list of target Twitter user IDs, exclusively used for searching recent tweets. 
     Each line in the uploaded file should represent a Twitter user ID.

- **Read File Content:**
    Retrieve the content of a specified file, providing each line in a proper format. 
    Ideal for accessing stored information or logs.

**Authentication:**
    
Access to these services is protected by an API key mechanism. Users must provide a valid API key in the request header for authentication.

**How to Use:**
    
- To upload main user cookies: Use the `/upload-reply-cookie/` endpoint with the appropriate API key.
    
- To upload scrap cookies: Utilize the `/upload-scrap-cookie/` endpoint with the correct API key.
    
- To upload target user IDs: Use the `/upload-target-ids/` endpoint, ensuring the provided API key is valid.

- To upload target user IDs as a DB excel: Use the `/upload-excel/` endpoint, providing an excel sheet of the accounts details and ensuring the provided API key is valid.

- To upload tweets replies: Use the `/tweets-replies-list/` endpoint, ensuring the provided API key is valid.

- To Check Target Account Details: Access the `/get-account/` specifying the account name and providing the API key for authentication.

- To display logs: Access the `/logs/` endpoint.
    
- To read file content: Access the `/read-file-content/` endpoint, specifying the file name and providing the API key for authentication.    



**Obtaining an API Key:**
    
For users requiring an API key, please contact `M Mansour` for assistance.

"""

app = FastAPI(title="Isnad Bot",
              description=description,
              summary="Isnad Scrap - Util API.",
              version="0.0.1", swagger_ui_parameters={"defaultModelsExpandDepth": -1}
              )

# SQLite database setup
DATABASE_URL = "sqlite:///./isnad9.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
Base = declarative_base()


class Account(Base):
    __tablename__ = "accounts"

    id = Column(Integer, primary_key=True, index=True)
    account_name = Column(String, index=True)
    account_id = Column(String)
    account_link = Column(String)
    account_status = Column(String)
    account_category = Column(String)
    account_type = Column(String)
    publishing_level = Column(String)
    access_level = Column(String)
    is_used = Column(Boolean, default=false())
    video_enabled = Column(Boolean, default=false())
    created_at = Column(DateTime, default=datetime.utcnow)


Base.metadata.create_all(bind=engine)

# Dependency to get the database session


def get_db():
    # Step 5: Insert data into the database
    Session = sessionmaker(bind=engine)
    session = Session()
    # db = session
    try:
        yield session
    finally:
        session.close()


def is_database_empty():
    """
    Check if the accounts table is empty.

    - **db**: Database session.

    Returns a boolean indicating whether the accounts table is empty.
    """
    Session = sessionmaker(bind=engine)
    session = Session()
    return not session.query(Account).first()


def check_database_status():
    """
    Check the status of the database and accounts table.

    - **db**: Database session.

    Returns a tuple (is_database_exist, is_table_exist, is_table_empty).
    """
    inspector = inspect(engine)
    is_database_exist = inspector.get_schema_names()

    if is_database_exist:
        is_table_exist = inspector.get_table_names()

        if is_table_exist and "accounts" in is_table_exist:
            is_table_empty = is_database_empty()
            return True, True, is_table_empty

    return False, False, False


# Variable to control the task execution
is_task_running = False

API_KEY_ADMIN = "iSLgvYQMFbExJGIVpJHEOEHnYxyzT4Fcr5xfSVG2Sn0q5FcrylK72Pgs3ctg0Cyp"

reply_cookie_file_path = 'twitter_reply_cookies.txt'
scrap_cookie_file_path = 'twitter_scrap_cookies.txt'
target_ids_file = 'target_user_ids.txt'


# Define a dictionary to store the mapping of userid to api_key
user_api_key_map = {
    "user1": "Hw1MXuWmKwsG4UXlRITVvS3vkKd5xvkiKD2Z9lXPvXZ5tuUEsTGAfqT8m8AnNGuo",
    "user2": "ERGZdjZqumtfZccYmpYwIlAO83RrqATioi6OUXQI8iiVZtG3xiKBfGgPjqgMwdvw",
    "user3": "WbpClFZ2HnsNgbYBwsoYeVFqUGYu64a71Thj7qHA9xE7ca8zjKFw1rOQzohwVOKX",
    "user4": "wRkBnrIMx20TPYRduKsq3SXfc8WkXh0Pj3H0hGYGJBT7qoXXxYMzTMk4JUqkyMsl",
    "user5": "AfIZUKMWVNo0KDnMdinHqaFIZnDgEWzDBw2PgubmffcQzUj9Lh5WaTz3ilzFx8Dp",
    "user6": "XQ5ihKJl2GqUvMM8O0Fs06UmZy6d0EeF4u3QLAMVbppzJETTul90PhQh7vI9oC4R",
    "user7": "fEpvjfO0oZr4fb3ncjQyAYdOc3DdkiCEhlKfBiNa8biHRHTly2duw20C44QZHBCf",
    "user8": "fKrFNeOwapEe7XrIiTRl9ufMbmxEaNGazYpemjb2VVkS8Z40fYVtgQMC46A26K7o",
    "user9": "wJkclhGS7ROCf13YncA69meOp7sdK5iAp9ofMYxbAUc9Gm7fFJ94Xj8EEnTqIEDq",
    "user10": "RpnNbdW9sN8zVddnuHrsFG0I2VMzY2VBI5tnrHhYqVrXiovvNsqHBNkpwjSwyyJf",
    "user2": "SedkImeAadVNu6TloPajo4ekQohXe5yWAi7aEb7aeeHVOeYqvI464Uj9cVXhjoVF",
    "user2": "onHGZ9U57aYeRWUiwoIpdRK4DVhCybGCQLO6xHXtW4SimPtas3j8f4K6OPpTVCEh",
    "admin": API_KEY_ADMIN
    # Add more users and their corresponding api keys
}


# Dependency to get the userid based on the provided api_key
async def get_api_key(api_key: str = Header(..., description="API key for authentication")):
    """
    Get User ID

    Retrieves the userid based on the provided API key.

    - **api_key**: API key for authentication.

    Returns the corresponding userid.
    """
    for userid, key in user_api_key_map.items():
        if key == api_key:
            return userid
    raise HTTPException(
        status_code=401,
        detail="Invalid API key",
        headers={"WWW-Authenticate": "Bearer"},
    )

# Dependency to check the admin API key


def get_admin_api_key(api_key: str = Header(..., description="Admin API key for authentication")):
    if api_key != API_KEY_ADMIN:
        raise HTTPException(
            status_code=401,
            detail="Invalid admin API key",
        )
    return api_key


@app.get("/", include_in_schema=False)
def read_root():
    return {"Isnad ": "Scrapper🔻🔻🔻"}


@app.on_event("startup")
async def startup_event():
    print('Node Scrapper Server started---- :', datetime.now())


# Define a function to check if the file is a text file
def is_text_file(file: UploadFile):
    if not file.content_type.startswith("text/"):
        raise HTTPException(
            status_code=400,
            detail="Only text files are allowed",
        )


@app.get("/read-file-content/", response_class=PlainTextResponse)
async def read_file_content(
    file_name: str = Query(..., title="File Name",
                           description="Name of the file to read."),
    api_key: str = Depends(get_api_key),
):
    """
    Read File Content

    Reads the content of the specified file and responds with each line in proper format.

    - **file_name**: Name of the file to read, including the extension, ex: .txt.
    - **api_key**: API key for authentication.

    Returns the content of the file as plain text.
    """
    try:
        with open(file_name, "r") as file:
            content = file.read()

        logger.info('Request from UserID: ' +
                    str(api_key) + '- read file '+file_name)
        return content
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"File '{file_name}' not found.",
        )

# reply_cookie_file_path = 'twitter_reply_cookies.txt'


@app.post("/upload-reply-cookie/")
async def upload_reply_cookie(
        file: UploadFile = File(...),
        api_key: str = Depends(get_api_key),
        replace_existing: bool = Query(default=False, description="Whether to replace existing content in twitter_reply_cookies.txt"), _: bool = Depends(is_text_file),):

    """
    Upload main user Cookies

    Update the list the main Twitter user cookies, who will reply/like the tweets.

    - **file**: List of main Twitter user cookies, the format should be auth_token_value|ct0_value|username, one record in each line. 
    - **api_key**: API key for authentication.
    - **replace_existing**: Whether to replace existing content in twitter_reply_cookies.txt.

    """

    contents = await file.read()

    # Decode contents and split it into lines
    lines = contents.decode("utf-8").splitlines()

    # Open file in append mode or write mode based on replace_existing flag
    mode = "w" if replace_existing else "a"

    # Append or replace content in the existing file "twitter_reply_cookies.txt"
    with open("twitter_reply_cookies.txt", mode, encoding="utf-8") as offline_file:
        for line in lines:
            if line.strip():  # Ignore empty lines
                offline_file.write(line + "\n")

    action = "Replaced" if replace_existing else "Appended"
    logger.info('Request from UserID: '+api_key +
                ' - File content '+action+' in twitter_reply_cookies.txt')
    return JSONResponse(content={"message": f"File content {action} in twitter_reply_cookies.txt"}, status_code=200)


# scrap_cookie_file_path = 'twitter_scrap_cookies.txt'
@app.post("/upload-scrap-cookie/")
async def upload_scrap_cookie(
        file: UploadFile = File(..., title="Scrap Twitter Accounts",
                                description="List of scrap Twitter user cookies, who will only be used for searching recent tweets."),
        api_key: str = Depends(get_api_key),
        replace_existing: bool = Query(default=False, description="Whether to replace existing content in twitter_scrap_cookies.txt"), _: bool = Depends(is_text_file),):

    """
    Upload scrap Cookies

    Update the list of scrap Twitter user cookies, who will only be used for searching recent tweets.

    - **file**: List of scrap Twitter user cookies, the format should be auth_token_value|ct0_value|username, one record in each line. 
    - **api_key**: API key for authentication.
    - **replace_existing**: Whether to replace existing content in twitter_scrap_cookies.txt.

    """

    contents = await file.read()

    # Decode contents and split it into lines
    lines = contents.decode("utf-8").splitlines()

    # Open file in append mode or write mode based on replace_existing flag
    mode = "w" if replace_existing else "a"

    # Append or replace content in the existing file "twitter_scrap_cookies.txt"
    with open("twitter_scrap_cookies.txt", mode, encoding="utf-8") as offline_file:
        for line in lines:
            if line.strip():  # Ignore empty lines
                offline_file.write(line + "\n")

    action = "Replaced" if replace_existing else "Appended"
    logger.info('Request from UserID: '+api_key +
                ' - File content '+action+' in twitter_scrap_cookies.txt')
    return JSONResponse(content={"message": f"File content {action} in twitter_scrap_cookies.txt"}, status_code=200)


# target_ids_file = 'target_user_ids.txt'
@app.post("/upload-target-ids/")
async def upload_target_ids(
        file: UploadFile = File(
            ..., description="Upload a text file containing reply cookies for Twitter."),
        api_key: str = Depends(get_api_key),
        replace_existing: bool = Query(default=False, description="Whether to replace existing content in target_user_ids.txt"), _: bool = Depends(is_text_file),):

    """
    Upload target ids

    Update the list of target Twitter Ids, who will only be used for searching recent tweets.

    - **file**: List of target Twitter user Ids, the format should be twitter user id in each line
    - **api_key**: API key for authentication.
    - **replace_existing**: Whether to replace existing content in target_user_ids.txt.

    """

    contents = await file.read()

    # Decode contents and split it into lines
    lines = contents.decode("utf-8").splitlines()

    # Open file in append mode or write mode based on replace_existing flag
    mode = "w" if replace_existing else "a"

    # Append or replace content in the existing file "target_user_ids.txt"
    with open("target_user_ids.txt", mode, encoding="utf-8") as offline_file:
        for line in lines:
            if line.strip():  # Ignore empty lines
                offline_file.write(line + "\n")

    action = "Replaced" if replace_existing else "Appended"
    logger.info('Request from UserID: '+api_key +
                ' - File content '+action+' in target_user_ids.txt')
    return JSONResponse(content={"message": f"File content {action} in target_user_ids.txt"}, status_code=200)


# target_ids_file = 'tweets_replies_list.txt'
@app.post("/tweets-replies-list/")
async def upload_tweets_replies(
        file: UploadFile = File(...,
                                description="Upload a text file containing the list of replies."),
        api_key: str = Depends(get_api_key),
        replace_existing: bool = Query(default=False, description="Whether to replace existing content in tweets_replies_list.txt"), _: bool = Depends(is_text_file),):

    """
    Upload tweets replies

    Update the list of replies, which will be used to reply to the recent tweets of targeted users.

    - **file**: List of tweets replies, the format should be one reply in each line
    - **api_key**: API key for authentication.
    - **replace_existing**: Whether to replace existing content in tweets_replies_list.txt.

    """

    contents = await file.read()

    # Decode contents and split it into lines
    lines = contents.decode("utf-8").splitlines()

    # Open file in append mode or write mode based on replace_existing flag
    mode = "w" if replace_existing else "a"

    # Append or replace content in the existing file "tweets_replies_list.txt"
    with open("tweets_replies_list.txt", mode, encoding="utf-8") as offline_file:
        for line in lines:
            if line.strip():  # Ignore empty lines
                offline_file.write(line + "\n")

    action = "Replaced" if replace_existing else "Appended"
    logger.info('Request from UserID: '+api_key +
                ' - File content '+action+' in tweets_replies_list.txt')
    return JSONResponse(content={"message": f"File content {action} in tweets_replies_list.txt"}, status_code=200)


# Dependency to check if the file is an Excel file
def is_excel_file(file: UploadFile):
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Only text files are allowed",
        )
    return file


# Endpoint to upload Excel file and add data to DB
@app.post("/upload-excel/")
async def upload_excel(
    api_key: str = Depends(get_api_key),
    file: UploadFile = Depends(is_excel_file),
    db: Session = Depends(get_db)
):
    """
    Upload Excel File and Add or Update Data in Database

    Allows the application to upload an Excel file, extract data, and add or update it in the database.

    - **file**: Upload an Excel file.
    - **db**: Database session.

    Returns a confirmation message.
    """
    try:
        # Load Excel file and extract data
        workbook = load_workbook(file.file)
        sheet = workbook.active

        # Map column names to indices
        header = sheet[1]
        column_indices = {header[i].value: i for i in range(len(header))}

        # Extract data from the Excel sheet and add or update it in the database
        for row in sheet.iter_rows(min_row=2, values_only=True):
            account_data = {
                "account_name": row[column_indices["ACCOUNT_NAME"]],
                "account_id": row[column_indices["ACCOUNT_ID"]],
                "account_link": row[column_indices["ACCOUNT_LINK"]],
                "account_status": row[column_indices["ACCOUNT_STATUS"]],
                "account_category": row[column_indices["ACCOUNT_CATEGORY"]],
                "account_type": row[column_indices["ACCOUNT_TYPE"]],
                "publishing_level": row[column_indices["PUBLISHING_LEVEL"]],
                "access_level": row[column_indices["ACCESS_LEVEL"]],
            }

            # Check if the record exists based on unique identifier (account_id)
            existing_record = db.query(Account).filter(Account.account_id == account_data["account_id"]).first()

            if existing_record:
                # Update existing record with new values
                for key, value in account_data.items():
                    setattr(existing_record, key, value)
            else:
                # Create a new record and add it to the database
                db_account = Account(**account_data)
                db.add(db_account)

        db.commit()
        logger.info('Request from UserID: ' +
                    api_key+' - Data added or updated in the database successfully')
        return JSONResponse(content={"message": "Data added or updated in the database successfully."}, status_code=200)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")


# Endpoint to get account details by ACCOUNT_NAME
@app.get("/get-account/")
async def get_account(
        api_key: str = Depends(get_api_key),
        account_name: str = Query(..., title="Account Id", description="The ACCOUNT_NAME to retrieve details for."), db: Session = Depends(get_db)):
    """
    Get Account Details by ACCOUNT_NAME

    Allows the application to retrieve account details from the database based on the provided ACCOUNT_NAME.

    - **account_name**: The ACCOUNT_NAME to retrieve details for.
    - **db**: Database session.

    Returns account details.
    """
    account = db.query(Account).filter(
        Account.account_name == account_name).first()

    if not account:
        raise HTTPException(
            status_code=404,
            detail="Account not found",
        )

    logger.info('Request from UserID: '+api_key +
                ' - Search for ACCOUNT_NAME'+account_name+' .')

    return {
        "account_name": account.account_name,
        "account_id": account.account_id,
        "account_link": account.account_link,
        "account_status": account.account_status,
        "account_category": account.account_category,
        "account_type": account.account_type,
        "publishing_level": account.publishing_level,
        "access_level": account.access_level,
        "video_enabled": account.video_enabled,
        "is_used": account.is_used,
        "created_at": account.created_at
    }


# Log viewer
@app.get("/logs", response_class=PlainTextResponse)
async def read_logs(api_key: str = Depends(get_api_key)):
    """
    View Application Logs

    Displays the application logs.

    Returns logs.
    """
    try:
        with open("app.log", "r", encoding="utf-8") as file:
            logs = file.readlines()

        # Reverse the order of logs to get the most recent entries first
        logs.reverse()

        content = "".join(logs)
        return content
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail=f"File app.log not found.",
        )


def start_scrap_background():
    global is_task_running
    is_task_running = True
    main()


@app.get("/start-scrap")
async def start_scrap(background_tasks: BackgroundTasks, api_key: str = Depends(get_admin_api_key)):
    background_tasks.add_task(start_scrap_background)
    return {"message": "Scrapper will start in the background..."}


@app.get("/stop-scrap")
async def start_scrap(background_tasks: BackgroundTasks, api_key: str = Depends(get_admin_api_key)):
    global is_task_running
    is_task_running = False
    return {"message": "Scrapper will stop now..."}


def get_time_difference_in_minutes(tweet_created_at):
    # Convert the tweet's creation time to a datetime object without timezone
    tweet_time_naive = datetime.strptime(
        tweet_created_at, '%a %b %d %H:%M:%S +0000 %Y').replace(tzinfo=None)

    # Set the timezone for the tweet's creation time to UTC
    tweet_time_utc = tweet_time_naive.replace(tzinfo=timezone.utc)

    # Get the current UTC time
    current_time_utc = datetime.now(timezone.utc)

    # Calculate the time difference
    time_difference = current_time_utc - tweet_time_utc

    # Calculate the time difference in minutes
    time_difference_minutes = time_difference.total_seconds() / 60

    return time_difference_minutes


class TweetEntry:
    def __init__(self, entry):
        self.rest_id = entry['content']['itemContent']['tweet_results']['result']['rest_id']
        self.created_at = entry['content']['itemContent']['tweet_results']['result']['legacy']['created_at']
        # Extract is_edit_eligible values
        is_edit_eligible1 = extract_is_edit_eligible(
            entry['content']['itemContent']['tweet_results']['result']['edit_control'])
        self.is_edit_eligible = is_edit_eligible1


def extract_is_edit_eligible(obj):
    if 'is_edit_eligible' in obj:
        return obj['is_edit_eligible']
    elif 'edit_control_initial' in obj and 'is_edit_eligible' in obj['edit_control_initial']:
        return obj['edit_control_initial']['is_edit_eligible']
    else:
        return None


# Initialize used_user_rest_ids as an empty set
used_user_rest_ids = set()


def get_user_last_tweets_file(scraper, TweetEntry):
    # Read user_rest_ids from a text file
    with open(target_ids_file, 'r', encoding='utf-8') as file:
        user_rest_ids = [line.strip() for line in file]

    # Shuffle the list to get a random order
    random.shuffle(user_rest_ids)

    latest_entries = []

    for user_rest_id in user_rest_ids:
        # Check if user_rest_id has already been used
        if user_rest_id not in used_user_rest_ids:
            try:
                tweetsScrap = scraper.tweets([user_rest_id])

                # Check for rate limit exceeded error
                if 'errors' in tweetsScrap and tweetsScrap['errors'][0]['code'] == 88:
                    print(
                        f"Rate limit exceeded. Unable to retrieve tweets for user_rest_id: {user_rest_id}")
                    break  # Stop the script when rate limit exceeded

                is_pin_included = len(
                    tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'])

                if is_pin_included == 3:
                    for entry in tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'][2]['entries']:
                        if entry['content']['entryType'] in {'TimelineTimelineItem'}:
                            tweet_entry = TweetEntry(entry)
                            if tweet_entry.is_edit_eligible:
                                latest_entries.append(tweet_entry)
                                # Mark user_rest_id as used
                                used_user_rest_ids.add(user_rest_id)
                                break  # Break the loop after processing one user_rest_id
                else:
                    for entry in tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'][1]['entries']:
                        if entry['content']['entryType'] in {'TimelineTimelineItem'}:
                            tweet_entry = TweetEntry(entry)
                            if tweet_entry.is_edit_eligible:
                                latest_entries.append(tweet_entry)
                                # Mark user_rest_id as used
                                used_user_rest_ids.add(user_rest_id)
                                break  # Break the loop after processing one user_rest_id
            except Exception as e:
                print(
                    f"An error occurred while processing user_rest_id {user_rest_id}: {str(e)}")

    return latest_entries


# Update the get_user_last_tweets function
def get_user_last_tweets_db(scraper, TweetEntry):
    Session = sessionmaker(bind=engine)
    session = Session()
    # Query for accounts that are not used
    accounts = (session.query(Account)
                .filter(Account.is_used == false())
                .order_by(Account.publishing_level, Account.access_level)
                .all()
                )

    # Check if the query returned no unused accounts
    if not accounts:
        # Reset is_used for all accounts to False and retry the query
        session.query(Account).update({Account.is_used: false()})
        session.commit()

        accounts = (session.query(Account)
                    .filter(Account.is_used == false())
                    .order_by(Account.publishing_level, Account.access_level)
                    .all()
                    )

    latest_entries = []
    for account in accounts:
        try:
            tweetsScrap = scraper.tweets([account.account_id])

            # Check for rate limit exceeded error
            if 'errors' in tweetsScrap and tweetsScrap['errors'][0]['code'] == 88:
                print(
                    f"Rate limit exceeded. Unable to retrieve tweets for account_id: {account.account_id}")
                break  # Stop the script when rate limit exceeded

            is_pin_included = len(
                tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'])

            if is_pin_included == 3:
                for entry in tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'][2]['entries']:
                    if entry['content']['entryType'] in {'TimelineTimelineItem'}:
                        tweet_entry = TweetEntry(entry)
                        if tweet_entry.is_edit_eligible:
                            latest_entries.append(tweet_entry)
                            # Mark account as used
                            account.is_used = True
                            break  # Break the loop after processing one account
            else:
                for entry in tweetsScrap[0]['data']['user']['result']['timeline_v2']['timeline']['instructions'][1]['entries']:
                    if entry['content']['entryType'] in {'TimelineTimelineItem'}:
                        tweet_entry = TweetEntry(entry)
                        if tweet_entry.is_edit_eligible:
                            latest_entries.append(tweet_entry)
                            # Mark account as used
                            account.is_used = True
                            break  # Break the loop after processing one account
        except Exception as e:
            print(
                f"An error occurred while processing account_id {account.account_id}: {str(e)}")

    # Commit changes to the database
    session.commit()

    return latest_entries

# Function to send tweets using auth_token_value, ct0_value, and plain_text


def send_reply(tweet_id):

    # Read tweets from a text file
    with open('tweets_replies_list.txt', 'r', encoding='utf-8') as file:
        replies = [line.strip() for line in file]

    # Shuffle the list to get a random order
    random.shuffle(replies)

    # Read cookies from the file
    with open(reply_cookie_file_path, 'r') as cookie_file:
        cookie_data = [line.strip().split('|') for line in cookie_file]

    # Shuffle the list to get a random order
    random.shuffle(cookie_data)

    hashtags = ["#BringThemHomeNow", "#Bringthemhome", "#WeWillNotStopUntilTheyAreAllBack",
                "#captured_oct7", "#Israel", "#hostages", "#MeTooUNlessYoureAJew",
                "#BringThemAllHomeNow", "#HamasTerrorists", "#BringThemHome",
                "#100Days", "#KfirBibas", "#KfirOneYear"]
    tweet_content = f" {random.choice(replies)}   {random.choice(hashtags)} {random.choice(hashtags)}"
    try:
        account = Account(
            cookies={"ct0": cookie_data[0][1], "auth_token": cookie_data[0][0]})
        account.reply(tweet_content, tweet_id=tweet_id)
        logger.info('User ' +
                    cookie_data[0][2] + ' finsihed reply on tweet: ' + str(tweet_id))
        print(f"{cookie_data[0][2]} finsihed reply on tweet {tweet_id}")
        pass
    except Exception as e:
        print(f"Exception in send_tweet Error: {e}")
        account = Account(
            cookies={"ct0": cookie_data[1][1], "auth_token": cookie_data[1][0]})
        account.reply(tweet_content, tweet_id=tweet_id)


def initialize_scraper() -> Scraper:
    # Read cookies from the file
    with open(scrap_cookie_file_path, 'r') as scrap_cookie_file:
        scrap_cookie_data = [line.strip().split('|')
                             for line in scrap_cookie_file]

    random.shuffle(scrap_cookie_data)
    # Loop over each pair of ct0 and auth_token
    for auth_token, ct0, user in scrap_cookie_data:
        try:
            scraper = Scraper(cookies={"ct0": ct0, "auth_token": auth_token})
            tweets = scraper.tweets_by_id([1749760475204554824])
            # Initialize a new scraper object in case of an error
            if 'errors' in tweets[0]:
                initialize_scraper()
            else:
                return scraper
        except Exception as e:
            print(f"Exception in initialize_scraper Error: {e}")


def main():
    # print('Run the function every hour')
    # Run the function every hour
    is_db_exist, is_table_exist, is_table_empty = check_database_status()
    while is_task_running:
        scraper = initialize_scraper()
        # Check the status of the database and accounts table
        is_db_exist, is_table_exist, is_table_empty = check_database_status()

        if not is_db_exist or not is_table_exist or not is_table_empty:
            latest_entries_list = get_user_last_tweets_db(scraper, TweetEntry)
        else:
            latest_entries_list = get_user_last_tweets_file(
                scraper, TweetEntry)
        # # Print details
        for index, entry in enumerate(latest_entries_list, start=1):
            # print(f"Entry {index}: rest_id: {entry.rest_id}, is_edit_eligible: {entry.is_edit_eligible} , created_at: {entry.created_at}")
            time_difference_minutes = get_time_difference_in_minutes(
                entry.created_at)
            print(f"Time difference: {time_difference_minutes}")
            # Tweet less than 5 minutes, means it's new, we will reply
            if time_difference_minutes < 5:
                print(
                    f"Entry {index}: rest_id: {entry.rest_id}, is_edit_eligible: {entry.is_edit_eligible} , created_at: {entry.created_at}")
                send_reply(entry.rest_id)
        time.sleep(3600)  # Sleep for 1 hour (3600 seconds)


if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
